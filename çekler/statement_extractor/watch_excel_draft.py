from __future__ import annotations

import argparse
import subprocess
import sys
import time
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from statement_extractor.src.excel_compare import build_default_diff_report_path, compare_excel_files, write_diff_report
from statement_extractor.src.snapshot_manager import create_snapshot, find_latest_snapshot
from statement_extractor.src.utils import format_try_amount

DEFAULT_EXCEL = ROOT / "outputs" / "jpeg_cek_taslak.xlsx"
DEFAULT_SNAPSHOTS_DIR = ROOT / "snapshots"
DEFAULT_DIFF_REPORTS_DIR = ROOT / "diff_reports"
SUMMARY_SHEET = "Doğrulama Özeti"
STATUS_HEADER = "Doğrulama Durumu"
NOTE_HEADER = "Doğrulama Notu"

OK_FILL = PatternFill("solid", fgColor="D9EAD3")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Taslak çek Excel dosyasını izler ve kaydedilince doğrular.")
    parser.add_argument("excel_path", nargs="?", default=str(DEFAULT_EXCEL), help="İzlenecek Excel dosyası")
    parser.add_argument("--sheet", help="Doğrulanacak sayfa adı. Verilmezse ilk veri sayfası kullanılır.")
    parser.add_argument("--interval", type=float, default=2.0, help="Dosya değişiklik kontrol aralığı")
    parser.add_argument("--once", action="store_true", help="Tek sefer doğrula ve çık")
    parser.add_argument("--no-open", action="store_true", help="Doğrulama sonrası Excel'i açma")
    parser.add_argument("--with-snapshot-diff", action="store_true", help="Kaydedilen Excel'i son snapshot ile karşılaştır")
    parser.add_argument("--snapshots-dir", default=str(DEFAULT_SNAPSHOTS_DIR), help="Snapshot klasörü")
    parser.add_argument("--diff-reports-dir", default=str(DEFAULT_DIFF_REPORTS_DIR), help="Diff rapor klasörü")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel_path)
    if not excel_path.exists():
        raise SystemExit(f"Excel dosyası bulunamadı: {excel_path}")

    if args.once:
        if args.with_snapshot_diff:
            run_snapshot_diff(excel_path, Path(args.snapshots_dir), Path(args.diff_reports_dir))
        validate_and_reopen(excel_path, sheet_name=args.sheet, open_after=not args.no_open)
        return 0

    print(f"İzleniyor: {excel_path}")
    print("Excel'i güncelleyip kaydettiğinde doğrulama yapılacak. Durdurmak için Ctrl+C.")
    last_mtime = excel_path.stat().st_mtime
    try:
        while True:
            time.sleep(args.interval)
            current_mtime = excel_path.stat().st_mtime
            if current_mtime == last_mtime:
                continue

            wait_until_file_is_stable(excel_path)
            if args.with_snapshot_diff:
                run_snapshot_diff(excel_path, Path(args.snapshots_dir), Path(args.diff_reports_dir))
            validate_and_reopen(excel_path, sheet_name=args.sheet, open_after=not args.no_open)
            last_mtime = excel_path.stat().st_mtime
    except KeyboardInterrupt:
        print("\nİzleme durduruldu.")
    return 0


def run_snapshot_diff(excel_path: Path, snapshots_dir: Path, diff_reports_dir: Path) -> None:
    try:
        latest_snapshot = find_latest_snapshot(excel_path.name, snapshots_dir)
        if latest_snapshot is None:
            create_snapshot(excel_path, snapshots_dir)
            print("İlk snapshot alındı.")
            return

        print(f"Son snapshot bulundu: {latest_snapshot.snapshot_file}")
        print("Yeni dosya ile karşılaştırılıyor...")
        report = compare_excel_files(
            latest_snapshot.snapshot_file,
            excel_path,
            old_snapshot_id=latest_snapshot.snapshot_id,
            new_snapshot_id=excel_path.stem,
            source_file=excel_path.name,
        )
        output_path = build_default_diff_report_path(excel_path.name, diff_reports_dir)
        write_diff_report(report, output_path)
        print(f"Değişen satır: {report.modified_count}")
        print(f"Eklenen satır: {report.added_count}")
        print(f"Silinen satır: {report.removed_count}")
        print(f"Tutar farkı: {format_try_amount(report.amount_difference)}")
        print(f"Diff raporu oluşturuldu: {output_path}")
        create_snapshot(excel_path, snapshots_dir)
        print("Yeni snapshot oluşturuldu.")
    except Exception as exc:
        print(f"Snapshot/diff hatası: {exc}")


def validate_and_reopen(excel_path: Path, sheet_name: str | None = None, open_after: bool = True) -> None:
    result = validate_workbook(excel_path, sheet_name=sheet_name)
    print(
        f"Doğrulandı: {result['total_rows']} satır, "
        f"{result['ok_count']} uygun, {result['warning_count']} kontrol, {result['error_count']} hatalı. "
        f"Toplam: {format_try(result['total_amount'])}"
    )
    if open_after:
        subprocess.run(["open", str(excel_path)], check=False)


def validate_workbook(excel_path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    workbook = load_workbook(excel_path)
    sheet = workbook[sheet_name] if sheet_name else find_data_sheet(workbook)
    headers = ensure_validation_headers(sheet)

    seen_keys: set[tuple[str, str, str, Decimal]] = set()
    total_rows = 0
    ok_count = 0
    warning_count = 0
    error_count = 0
    total_amount = Decimal("0")

    for row_number in range(2, sheet.max_row + 1):
        row_values = row_as_dict(sheet, headers, row_number)
        if is_empty_row(row_values) or is_total_row(row_values):
            continue

        total_rows += 1
        status, notes, amount, normalized_date = validate_row(row_values, seen_keys)
        if amount is not None:
            total_amount += amount

        if status == "UYGUN":
            ok_count += 1
            fill = OK_FILL
        elif status == "KONTROL":
            warning_count += 1
            fill = WARNING_FILL
        else:
            error_count += 1
            fill = ERROR_FILL

        status_col = headers[STATUS_HEADER]
        note_col = headers[NOTE_HEADER]
        sheet.cell(row=row_number, column=status_col).value = status
        sheet.cell(row=row_number, column=note_col).value = "; ".join(notes)
        apply_row_fill(sheet, row_number, fill)

        if normalized_date is not None and "Vade Tarihi" in headers:
            cell = sheet.cell(row=row_number, column=headers["Vade Tarihi"])
            cell.value = normalized_date
            cell.number_format = "dd.mm.yyyy"
        if amount is not None and "Tutar" in headers:
            cell = sheet.cell(row=row_number, column=headers["Tutar"])
            cell.value = float(amount)
            cell.number_format = '#,##0.00 "TL"'

    update_summary_sheet(workbook, total_rows, ok_count, warning_count, error_count, total_amount)
    workbook.save(excel_path)
    return {
        "total_rows": total_rows,
        "ok_count": ok_count,
        "warning_count": warning_count,
        "error_count": error_count,
        "total_amount": total_amount,
    }


def find_data_sheet(workbook: Any) -> Any:
    for sheet in workbook.worksheets:
        if sheet.title != SUMMARY_SHEET:
            return sheet
    return workbook.active


def ensure_validation_headers(sheet: Any) -> dict[str, int]:
    headers = {
        str(sheet.cell(row=1, column=column).value).strip(): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=column).value
    }

    for header in (STATUS_HEADER, NOTE_HEADER):
        if header not in headers:
            column = sheet.max_column + 1
            sheet.cell(row=1, column=column).value = header
            sheet.cell(row=1, column=column).fill = HEADER_FILL
            sheet.cell(row=1, column=column).font = HEADER_FONT
            headers[header] = column
    return headers


def row_as_dict(sheet: Any, headers: dict[str, int], row_number: int) -> dict[str, Any]:
    return {header: sheet.cell(row=row_number, column=column).value for header, column in headers.items()}


def validate_row(
    row: dict[str, Any],
    seen_keys: set[tuple[str, str, str, Decimal]],
) -> tuple[str, list[str], Decimal | None, date | None]:
    notes: list[str] = []
    check_no = normalize_text(row.get("Çek No"))
    bank = normalize_text(row.get("Banka"))
    currency = normalize_text(row.get("Para Birimi")) or "TRY"
    amount = parse_amount(row.get("Tutar"))
    maturity_date = parse_date(row.get("Vade Tarihi"))

    if not check_no:
        notes.append("Çek no boş.")
    elif "?" in check_no:
        notes.append("Çek no net değil, kontrol edilmeli.")

    if not bank:
        notes.append("Banka boş.")
    if maturity_date is None:
        notes.append("Vade tarihi okunamadı.")
    if amount is None:
        notes.append("Tutar okunamadı.")
    elif amount <= 0:
        notes.append("Tutar pozitif olmalı.")
    if currency and currency.upper() != "TRY":
        notes.append("Para birimi TRY değil, kontrol edilmeli.")

    if check_no and bank and maturity_date and amount is not None:
        key = (check_no, bank.upper(), maturity_date.isoformat(), amount)
        if key in seen_keys:
            notes.append("Aynı çek no + banka + vade + tutar tekrar ediyor.")
        else:
            seen_keys.add(key)

    if not notes:
        return "UYGUN", ["Satır doğrulandı."], amount, maturity_date
    if any("boş" in note or "okunamadı" in note or "pozitif" in note for note in notes):
        return "HATALI", notes, amount, maturity_date
    return "KONTROL", notes, amount, maturity_date


def parse_amount(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).replace("TL", "").replace("₺", "").strip()
    text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            if fmt == "%d.%m.%y" and parsed.year < 2000:
                return parsed.replace(year=parsed.year + 100)
            return parsed
        except ValueError:
            continue
    return None


def update_summary_sheet(
    workbook: Any,
    total_rows: int,
    ok_count: int,
    warning_count: int,
    error_count: int,
    total_amount: Decimal,
) -> None:
    if SUMMARY_SHEET in workbook.sheetnames:
        del workbook[SUMMARY_SHEET]
    sheet = workbook.create_sheet(SUMMARY_SHEET)
    rows = [
        ("Doğrulama zamanı", datetime.now().strftime("%d.%m.%Y %H:%M:%S")),
        ("Toplam satır", total_rows),
        ("Uygun satır", ok_count),
        ("Kontrol gerektiren satır", warning_count),
        ("Hatalı satır", error_count),
        ("Toplam tutar", float(total_amount)),
    ]
    for row in rows:
        sheet.append(row)
    for row_index in range(1, sheet.max_row + 1):
        sheet.cell(row=row_index, column=1).font = Font(bold=True)
    sheet.cell(row=6, column=2).number_format = '#,##0.00 "TL"'
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 24


def apply_row_fill(sheet: Any, row_number: int, fill: PatternFill) -> None:
    for cell in sheet[row_number]:
        cell.fill = fill


def is_empty_row(row: dict[str, Any]) -> bool:
    data_headers = ["Banka", "Vade Tarihi", "Tutar", "Çek No"]
    return all(not row.get(header) for header in data_headers)


def is_total_row(row: dict[str, Any]) -> bool:
    value = row.get("Sıra")
    return normalize_text(value).upper() == "TOPLAM"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def format_try(value: Decimal) -> str:
    formatted = f"{value:,.2f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".") + " TL"


def wait_until_file_is_stable(path: Path, checks: int = 3, delay: float = 0.4) -> None:
    previous_size = -1
    stable_count = 0
    while stable_count < checks:
        current_size = path.stat().st_size
        if current_size == previous_size:
            stable_count += 1
        else:
            stable_count = 0
            previous_size = current_size
        time.sleep(delay)


if __name__ == "__main__":
    raise SystemExit(main())
