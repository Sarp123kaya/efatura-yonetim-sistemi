from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .diff_models import DiffType, ExcelDiffReport, FieldChange, RowDiff
from .utils import safe_filename


COMPARE_FIELDS = [
    "Çek No",
    "Banka",
    "Vade Tarihi",
    "Vadeye Kalan Gün",
    "Tutar",
    "Lehtar",
    "Keşideci",
    "Gönderildiği Yer",
    "Açıklama",
    "Doğrulama Durumu",
    "Doğrulama Notu",
]
SHEET_SKIP_NAMES = {"Özet", "Banka Özeti", "Vade Takvimi", "Değer Hesabı", "Parse Debug", "Doğrulama Özeti"}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ADDED_FILL = PatternFill("solid", fgColor="D9EAD3")
REMOVED_FILL = PatternFill("solid", fgColor="F4CCCC")
MODIFIED_FILL = PatternFill("solid", fgColor="FFF2CC")
TRY_FORMAT = '#,##0.00 "TL"'
DATE_FORMAT = "dd.mm.yyyy"


def compare_excel_files(
    old_excel_path: str | Path,
    new_excel_path: str | Path,
    old_snapshot_id: Optional[str] = None,
    new_snapshot_id: Optional[str] = None,
    source_file: Optional[str] = None,
) -> ExcelDiffReport:
    old_data = load_excel_rows(old_excel_path)
    new_data = load_excel_rows(new_excel_path)
    old_rows = old_data["rows"]
    new_rows = new_data["rows"]

    old_index = {row["row_identity"]: row for row in old_rows}
    new_index = {row["row_identity"]: row for row in new_rows}
    row_diffs: list[RowDiff] = []

    for identity in sorted(set(old_index) | set(new_index)):
        old_row = old_index.get(identity)
        new_row = new_index.get(identity)
        if old_row is None and new_row is not None:
            row_diffs.append(build_row_diff(DiffType.ADDED, identity, None, new_row))
        elif new_row is None and old_row is not None:
            row_diffs.append(build_row_diff(DiffType.REMOVED, identity, old_row, None))
        elif old_row is not None and new_row is not None:
            changes = compare_rows(identity, old_row, new_row)
            diff_type = DiffType.MODIFIED if changes else DiffType.UNCHANGED
            row_diffs.append(build_row_diff(diff_type, identity, old_row, new_row, changes))

    old_total = sum((row.get("amount_decimal") or Decimal("0") for row in old_rows), Decimal("0"))
    new_total = sum((row.get("amount_decimal") or Decimal("0") for row in new_rows), Decimal("0"))
    return ExcelDiffReport(
        old_snapshot_id=old_snapshot_id or Path(old_excel_path).stem,
        new_snapshot_id=new_snapshot_id or Path(new_excel_path).stem,
        source_file=source_file or Path(new_excel_path).name,
        generated_at=datetime.now(),
        account_code=new_data.get("account_code") or old_data.get("account_code"),
        account_name=new_data.get("account_name") or old_data.get("account_name"),
        company_name=new_data.get("company_name") or old_data.get("company_name"),
        added_count=count_type(row_diffs, DiffType.ADDED),
        removed_count=count_type(row_diffs, DiffType.REMOVED),
        modified_count=count_type(row_diffs, DiffType.MODIFIED),
        unchanged_count=count_type(row_diffs, DiffType.UNCHANGED),
        old_total_amount=old_total,
        new_total_amount=new_total,
        amount_difference=new_total - old_total,
        row_diffs=row_diffs,
    )


def load_excel_rows(excel_path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = find_main_sheet(workbook)
    headers = read_headers(sheet)
    rows: list[dict[str, Any]] = []
    used_identities: dict[str, int] = {}

    for row_index in range(2, sheet.max_row + 1):
        raw_row = {header: sheet.cell(row=row_index, column=column).value for header, column in headers.items()}
        if is_empty_or_total_row(raw_row):
            continue
        normalized = normalize_row(raw_row, row_index)
        identity = normalized["row_identity"]
        used_count = used_identities.get(identity, 0)
        used_identities[identity] = used_count + 1
        if used_count:
            normalized["row_identity"] = f"{identity}#dup{used_count + 1}"
        rows.append(normalized)

    metadata = extract_metadata(workbook)
    metadata.update({"rows": rows, "sheet_name": sheet.title})
    return metadata


def find_main_sheet(workbook: Any) -> Any:
    for name in ("Çek Listesi", "JPEG Çek Taslak"):
        if name in workbook.sheetnames:
            return workbook[name]
    for sheet in workbook.worksheets:
        if sheet.title not in SHEET_SKIP_NAMES:
            return sheet
    return workbook.active


def read_headers(sheet: Any) -> dict[str, int]:
    return {
        str(sheet.cell(row=1, column=column).value).strip(): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=column).value
    }


def normalize_row(raw_row: dict[str, Any], row_index: int) -> dict[str, Any]:
    row = {
        "row_index": row_index,
        "Çek No": text_value(get_alias(raw_row, "Çek No")),
        "Banka": text_value(get_alias(raw_row, "Banka")),
        "Vade Tarihi": normalize_date(get_alias(raw_row, "Vade Tarihi")),
        "Vadeye Kalan Gün": normalize_integer(get_alias(raw_row, "Vadeye Kalan Gün")),
        "Tutar": normalize_amount(get_alias(raw_row, "Tutar")),
        "Lehtar": text_value(get_alias(raw_row, "Lehtar", "Lehtar / Alacaklı", "customer_name")),
        "Keşideci": text_value(get_alias(raw_row, "Keşideci", "Keşideci / Firma", "company_name")),
        "Gönderildiği Yer": text_value(get_alias(raw_row, "Gönderildiği Yer", "sent_to")),
        "Açıklama": text_value(get_alias(raw_row, "Açıklama", "raw_description")),
        "Doğrulama Durumu": text_value(get_alias(raw_row, "Doğrulama Durumu")),
        "Doğrulama Notu": text_value(get_alias(raw_row, "Doğrulama Notu")),
    }
    row["amount_decimal"] = parse_amount(row["Tutar"])
    row["maturity_date"] = parse_date(row["Vade Tarihi"])
    identity, method = build_identity(row, row_index)
    row["row_identity"] = identity
    row["identity_method"] = method
    return row


def build_identity(row: dict[str, Any], row_index: int) -> tuple[str, str]:
    check_no = row.get("Çek No")
    if check_no:
        return f"check_no:{check_no}", "check_no"
    bank = row.get("Banka")
    maturity = row.get("maturity_date")
    amount = row.get("amount_decimal")
    if bank and maturity and amount is not None:
        return f"bank_maturity_amount:{bank}|{maturity.isoformat()}|{amount}", "bank_maturity_amount"
    return f"row_index:{row_index}", "row_index"


def compare_rows(identity: str, old_row: dict[str, Any], new_row: dict[str, Any]) -> list[FieldChange]:
    changes: list[FieldChange] = []
    for field in COMPARE_FIELDS:
        old_value = old_row.get(field)
        new_value = new_row.get(field)
        if old_value != new_value:
            changes.append(FieldChange(identity, field, old_value, new_value))
    return changes


def build_row_diff(
    diff_type: DiffType,
    identity: str,
    old_row: Optional[dict[str, Any]],
    new_row: Optional[dict[str, Any]],
    changes: Optional[list[FieldChange]] = None,
) -> RowDiff:
    row = new_row or old_row or {}
    return RowDiff(
        diff_type=diff_type,
        row_identity=identity,
        check_no=row.get("Çek No"),
        bank=row.get("Banka"),
        maturity_date=row.get("maturity_date"),
        amount=row.get("amount_decimal"),
        changes=changes or [],
        identity_method=row.get("identity_method", "row_index"),
        old_row=old_row or {},
        new_row=new_row or {},
    )


def write_diff_report(report: ExcelDiffReport, output_path: str | Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_summary_sheet(workbook, report)
    add_changed_sheet(workbook, report)
    add_row_sheet(workbook, "Eklenen Satırlar", [row for row in report.row_diffs if row.diff_type == DiffType.ADDED])
    add_row_sheet(workbook, "Silinen Satırlar", [row for row in report.row_diffs if row.diff_type == DiffType.REMOVED])
    add_all_comparison_sheet(workbook, report)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def build_default_diff_report_path(source_file: str, diff_reports_dir: str | Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    source_name = safe_filename(Path(source_file).stem, default="excel")
    return Path(diff_reports_dir) / f"diff_report_{source_name}_{timestamp}.xlsx"


def add_summary_sheet(workbook: Workbook, report: ExcelDiffReport) -> None:
    sheet = workbook.create_sheet("Fark Özeti")
    rows = [
        ("Kaynak dosya", report.source_file),
        ("Eski snapshot", report.old_snapshot_id),
        ("Yeni snapshot", report.new_snapshot_id),
        ("Cari hesap kodu", report.account_code),
        ("Cari hesap adı", report.account_name),
        ("Firma adı", report.company_name),
        ("Eski toplam tutar", float(report.old_total_amount)),
        ("Yeni toplam tutar", float(report.new_total_amount)),
        ("Tutar farkı", float(report.amount_difference)),
        ("Eklenen satır sayısı", report.added_count),
        ("Silinen satır sayısı", report.removed_count),
        ("Değişen satır sayısı", report.modified_count),
        ("Değişmeyen satır sayısı", report.unchanged_count),
    ]
    for row in rows:
        sheet.append(row)
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=1).font = Font(bold=True)
    for row in (7, 8, 9):
        sheet.cell(row=row, column=2).number_format = TRY_FORMAT
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 42


def add_changed_sheet(workbook: Workbook, report: ExcelDiffReport) -> None:
    sheet = workbook.create_sheet("Değişen Satırlar")
    headers = ["Satır identity", "Identity Yöntemi", "Çek No", "Banka", "Vade Tarihi", "Tutar", "Değişen Alan", "Eski Değer", "Yeni Değer"]
    append_header(sheet, headers)
    for row_diff in report.row_diffs:
        if row_diff.diff_type != DiffType.MODIFIED:
            continue
        for change in row_diff.changes:
            sheet.append(
                [
                    row_diff.row_identity,
                    row_diff.identity_method,
                    row_diff.check_no,
                    row_diff.bank,
                    row_diff.maturity_date,
                    float(row_diff.amount) if row_diff.amount is not None else None,
                    change.field_name,
                    change.old_value,
                    change.new_value,
                ]
            )
    format_sheet(sheet)
    format_columns(sheet, date_columns=[5], amount_columns=[6])


def add_row_sheet(workbook: Workbook, title: str, rows: list[RowDiff]) -> None:
    sheet = workbook.create_sheet(title)
    headers = ["Satır identity", "Identity Yöntemi", "Çek No", "Banka", "Vade Tarihi", "Vadeye Kalan Gün", "Tutar", "Lehtar", "Keşideci", "Açıklama", "Doğrulama Durumu", "Doğrulama Notu"]
    append_header(sheet, headers)
    for row_diff in rows:
        row = row_diff.new_row if row_diff.diff_type == DiffType.ADDED else row_diff.old_row
        sheet.append(row_to_sheet_values(row_diff, row))
    format_sheet(sheet)
    format_columns(sheet, date_columns=[5], amount_columns=[7])


def add_all_comparison_sheet(workbook: Workbook, report: ExcelDiffReport) -> None:
    sheet = workbook.create_sheet("Tüm Karşılaştırma")
    headers = ["Diff Tipi", "Satır identity", "Identity Yöntemi", "Çek No", "Banka", "Vade Tarihi", "Tutar", "Değişiklik Özeti"]
    append_header(sheet, headers)
    for row_diff in report.row_diffs:
        sheet.append(
            [
                row_diff.diff_type.value,
                row_diff.row_identity,
                row_diff.identity_method,
                row_diff.check_no,
                row_diff.bank,
                row_diff.maturity_date,
                float(row_diff.amount) if row_diff.amount is not None else None,
                summarize_changes(row_diff),
            ]
        )
        fill = None
        if row_diff.diff_type == DiffType.ADDED:
            fill = ADDED_FILL
        elif row_diff.diff_type == DiffType.REMOVED:
            fill = REMOVED_FILL
        elif row_diff.diff_type == DiffType.MODIFIED:
            fill = MODIFIED_FILL
        if fill:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill
    format_sheet(sheet)
    format_columns(sheet, date_columns=[6], amount_columns=[7])


def row_to_sheet_values(row_diff: RowDiff, row: dict[str, Any]) -> list[Any]:
    return [
        row_diff.row_identity,
        row_diff.identity_method,
        row.get("Çek No"),
        row.get("Banka"),
        row.get("maturity_date"),
        row.get("Vadeye Kalan Gün"),
        float(row.get("amount_decimal")) if row.get("amount_decimal") is not None else None,
        row.get("Lehtar"),
        row.get("Keşideci"),
        row.get("Açıklama"),
        row.get("Doğrulama Durumu"),
        row.get("Doğrulama Notu"),
    ]


def summarize_changes(row_diff: RowDiff) -> str:
    if not row_diff.changes:
        return ""
    return "; ".join(f"{change.field_name}: {change.old_value} -> {change.new_value}" for change in row_diff.changes)


def extract_metadata(workbook: Any) -> dict[str, Any]:
    metadata = {"account_code": None, "account_name": None, "company_name": None}
    if "Özet" not in workbook.sheetnames:
        return metadata
    sheet = workbook["Özet"]
    labels = {
        "Alt hesap kodu": "account_code",
        "Alt hesap adı": "account_name",
        "Firma adı": "company_name",
    }
    for row in sheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = labels.get(str(row[0]).strip())
        if key:
            metadata[key] = row[1] if len(row) > 1 else None
    return metadata


def get_alias(row: dict[str, Any], *aliases: str) -> Any:
    for alias in aliases:
        if alias in row:
            return row.get(alias)
    return None


def is_empty_or_total_row(row: dict[str, Any]) -> bool:
    if str(row.get("Sıra") or "").strip().upper() == "TOPLAM":
        return True
    return not any(row.get(header) for header in ("Çek No", "Banka", "Vade Tarihi", "Tutar"))


def text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_date(value: Any) -> str:
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else text_value(value)


def normalize_amount(value: Any) -> str:
    amount = parse_amount(value)
    return str(amount) if amount is not None else text_value(value)


def normalize_integer(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return str(int(value))
    except (TypeError, ValueError):
        return text_value(value)


def parse_amount(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = str(value).replace("TL", "").replace("₺", "").strip()
    text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d.%m.%y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            if fmt == "%d.%m.%y" and parsed.year < 2000:
                return parsed.replace(year=parsed.year + 100)
            return parsed
        except ValueError:
            continue
    return None


def count_type(rows: list[RowDiff], diff_type: DiffType) -> int:
    return sum(1 for row in rows if row.diff_type == diff_type)


def append_header(sheet: Any, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def format_sheet(sheet: Any) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 60)
    if sheet.max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions


def format_columns(sheet: Any, date_columns: list[int], amount_columns: list[int]) -> None:
    for column in date_columns:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column).number_format = DATE_FORMAT
    for column in amount_columns:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column).number_format = TRY_FORMAT
