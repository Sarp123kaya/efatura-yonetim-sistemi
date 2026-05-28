from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_writer import calculate_weighted_average_maturity
from .models import StoredCheckRecord
from .utils import calculate_days_until_maturity, calculate_maturity_status, get_week_period


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9EAD3")
TRY_FORMAT = '#,##0.00 "TL"'
DATE_FORMAT = "dd.mm.yyyy"


def write_consolidated_check_report(
    records: list[StoredCheckRecord],
    output_path: str | Path,
    today: date | None = None,
    include_overdue_in_value_calc: bool = False,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sorted_records = sorted(records, key=lambda item: (item.maturity_date, item.amount or Decimal("0"), item.check_no))
    add_all_checks_sheet(workbook, sorted_records, today)
    add_account_summary_sheet(workbook, sorted_records, today)
    add_bank_summary_sheet(workbook, sorted_records)
    add_maturity_calendar_sheet(workbook, sorted_records)
    add_value_sheet(workbook, sorted_records, today, include_overdue_in_value_calc)
    add_source_history_sheet(workbook, sorted_records)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def add_all_checks_sheet(workbook: Workbook, records: list[StoredCheckRecord], today: date | None) -> None:
    sheet = workbook.create_sheet("Tüm Çekler")
    headers = [
        "Sıra",
        "Cari / Lehtar",
        "Keşideci / Firma",
        "Çek No",
        "Banka",
        "Vade Tarihi",
        "Vade Ayı",
        "Vadeye Kalan Gün",
        "Vade Durumu",
        "Tutar",
        "Para Birimi",
        "Gönderildiği Yer",
        "Kaynak Tipi",
        "Kaynak Dosya",
        "Kontrol Durumu",
        "Import Session",
        "Açıklama",
        "Parse Uyarısı",
    ]
    append_header(sheet, headers)
    total = Decimal("0")
    for index, record in enumerate(records, start=1):
        total += record.amount or Decimal("0")
        sheet.append(
            [
                index,
                record.customer_name,
                record.company_name,
                record.check_no,
                record.bank,
                record.maturity_date,
                record.maturity_month,
                calculate_days_until_maturity(record.maturity_date, today),
                calculate_maturity_status(record.maturity_date, today),
                float(record.amount or Decimal("0")),
                record.currency,
                record.sent_to,
                record.source_type.value,
                record.source_file,
                record.review_status.value,
                record.import_session_id,
                record.raw_description,
                record.parse_warning,
            ]
        )
    sheet.append(["Toplam"] + [""] * 8 + [float(total)] + [""] * 8)
    fill_row(sheet, sheet.max_row, TOTAL_FILL)
    format_sheet(sheet)
    format_columns(sheet, date_columns=[6], amount_columns=[10])


def add_account_summary_sheet(workbook: Workbook, records: list[StoredCheckRecord], today: date | None) -> None:
    sheet = workbook.create_sheet("Cari Bazlı Özet")
    append_header(
        sheet,
        [
            "Cari / Lehtar",
            "Çek adedi",
            "Toplam tutar",
            "Ödenen çek adedi",
            "Ödenen tutar",
            "En erken vade",
            "En geç vade",
        ],
    )
    grouped: dict[str, list[StoredCheckRecord]] = defaultdict(list)
    for record in records:
        grouped[record.customer_name or "Belirsiz Cari"].append(record)
    for account, items in sorted(grouped.items()):
        total = sum((item.amount or Decimal("0") for item in items), Decimal("0"))
        paid_items = [item for item in items if is_paid_check(item, today)]
        paid_total = sum((item.amount or Decimal("0") for item in paid_items), Decimal("0"))
        maturities = [item.maturity_date for item in items]
        sheet.append([account, len(items), float(total), len(paid_items), float(paid_total), min(maturities), max(maturities)])
    format_sheet(sheet)
    format_columns(sheet, date_columns=[6, 7], amount_columns=[3, 5])


def is_paid_check(record: StoredCheckRecord, today: date | None) -> bool:
    if today is None:
        today = date.today()
    return record.maturity_date < today


def add_bank_summary_sheet(workbook: Workbook, records: list[StoredCheckRecord]) -> None:
    sheet = workbook.create_sheet("Banka Özeti")
    append_header(sheet, ["Banka", "Çek adedi", "Toplam tutar", "Ortalama tutar", "En erken vade", "En geç vade"])
    grouped: dict[str, list[StoredCheckRecord]] = defaultdict(list)
    for record in records:
        grouped[record.bank].append(record)
    for bank, items in sorted(grouped.items()):
        total = sum((item.amount or Decimal("0") for item in items), Decimal("0"))
        maturities = [item.maturity_date for item in items]
        sheet.append([bank, len(items), float(total), float(total / len(items)), min(maturities), max(maturities)])
    format_sheet(sheet)
    format_columns(sheet, date_columns=[5, 6], amount_columns=[3, 4])


def add_maturity_calendar_sheet(workbook: Workbook, records: list[StoredCheckRecord]) -> None:
    sheet = workbook.create_sheet("Vade Takvimi")
    append_header(sheet, ["Vade Ayı", "Adet", "Toplam"])
    months: dict[str, list[StoredCheckRecord]] = defaultdict(list)
    weeks: dict[str, list[StoredCheckRecord]] = defaultdict(list)
    for record in records:
        months[record.maturity_month].append(record)
        weeks[get_week_period(record.maturity_date)].append(record)
    for month, items in sorted(months.items()):
        total = sum((item.amount or Decimal("0") for item in items), Decimal("0"))
        sheet.append([month, len(items), float(total)])
    sheet.append([])
    append_header(sheet, ["Haftalık Vade Periyodu", "Adet", "Toplam"])
    for week, items in sorted(weeks.items()):
        total = sum((item.amount or Decimal("0") for item in items), Decimal("0"))
        sheet.append([week, len(items), float(total)])
    format_sheet(sheet)
    format_columns(sheet, date_columns=[], amount_columns=[3])


def add_value_sheet(
    workbook: Workbook,
    records: list[StoredCheckRecord],
    today: date | None,
    include_overdue: bool,
) -> None:
    sheet = workbook.create_sheet("Değer Hesabı")
    append_header(sheet, ["Sıra", "Çek No", "Banka", "Vade Tarihi", "Vadeye Kalan Gün", "Tutar", "Gün × Tutar"])
    total_amount = Decimal("0")
    total_weighted = Decimal("0")
    for index, record in enumerate(records, start=1):
        amount = record.amount or Decimal("0")
        days = calculate_days_until_maturity(record.maturity_date, today)
        if days < 0 and not include_overdue:
            continue
        weighted = Decimal(days) * amount
        total_amount += amount
        total_weighted += weighted
        sheet.append([index, record.check_no, record.bank, record.maturity_date, days, float(amount), float(weighted)])
    average = calculate_weighted_average_maturity(records, today=today, include_overdue=include_overdue)
    sheet.append([])
    sheet.append(["Toplam Tutar", "", "", "", "", float(total_amount), ""])
    sheet.append(["Toplam Gün × Tutar", "", "", "", "", "", float(total_weighted)])
    sheet.append(["Ağırlıklı Ortalama Vade", "", "", "", f"{average:.2f} gün", "", ""])
    if not include_overdue:
        sheet.append(["Not", "Vadesi geçmiş çekler ağırlıklı ortalama vade hesabına dahil edilmemiştir.", "", "", "", "", ""])
    format_sheet(sheet)
    format_columns(sheet, date_columns=[4], amount_columns=[6])


def add_source_history_sheet(workbook: Workbook, records: list[StoredCheckRecord]) -> None:
    sheet = workbook.create_sheet("Kaynak ve Kontrol Geçmişi")
    append_header(sheet, ["Kaynak Tipi", "Kaynak Dosya", "Kontrol Durumu", "Import Session", "Çek No", "Banka", "Vade", "Tutar", "Uyarı"])
    for record in records:
        sheet.append(
            [
                record.source_type.value,
                record.source_file,
                record.review_status.value,
                record.import_session_id,
                record.check_no,
                record.bank,
                record.maturity_date,
                float(record.amount or Decimal("0")),
                record.parse_warning,
            ]
        )
    format_sheet(sheet)
    format_columns(sheet, date_columns=[7], amount_columns=[8])


def build_default_consolidated_output_path(output_dir: str | Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(output_dir) / f"consolidated_checks_{timestamp}.xlsx"


def append_header(sheet: Any, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def fill_row(sheet: Any, row_number: int, fill: PatternFill) -> None:
    for cell in sheet[row_number]:
        cell.fill = fill


def format_sheet(sheet: Any) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)
    if sheet.max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions


def format_columns(sheet: Any, date_columns: list[int], amount_columns: list[int]) -> None:
    for column in date_columns:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column).number_format = DATE_FORMAT
    for column in amount_columns:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column).number_format = TRY_FORMAT
