from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import StatementMetadata
from .parser import is_received_check_line, parse_received_check_line
from .utils import format_try_amount, parse_turkish_date
from .zirve_muavin_excel import (
    build_synthetic_muavin_line,
    find_muavin_header_row,
    parse_zirve_banner,
    pick_movement_amount,
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9EAD3")
TRY_FORMAT = '#,##0.00 "TL"'
DATE_FORMAT = "dd.mm.yyyy"
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


@dataclass
class CustomerMovement:
    customer_name: str
    account_code: str | None
    source_file: str
    source_sheet: str
    source_row_index: int
    transaction_date: date | None
    voucher_no: str | None
    document_date: date | None
    document_no: str | None
    movement_type: str
    description: str | None
    debit: Decimal
    credit: Decimal
    balance: Decimal
    balance_type: str | None


@dataclass
class CustomerCheck:
    customer_name: str
    account_code: str | None
    source_file: str
    source_row_index: int
    transaction_date: date | None
    check_no: str
    bank: str
    maturity_date: date
    amount: Decimal
    description: str
    warning: str | None = None


@dataclass
class CustomerReportResult:
    output_path: Path
    file_count: int
    customer_count: int
    movement_count: int
    check_count: int
    check_total: Decimal


def parse_customer_statement_excels(input_dir: str | Path) -> tuple[list[CustomerMovement], list[CustomerCheck], int]:
    """Parse Zirve muavin Excel uploads into movements and received-check rows."""
    movements: list[CustomerMovement] = []
    checks: list[CustomerCheck] = []
    files = [path for path in sorted(Path(input_dir).iterdir()) if path.suffix.lower() in SUPPORTED_EXCEL_SUFFIXES]

    for excel_path in files:
        workbook = load_workbook(excel_path, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                header = find_muavin_header_row(sheet)
                if not header:
                    continue

                header_row, colmap = header
                banner = parse_zirve_banner(sheet, header_row)
                metadata = StatementMetadata(
                    company_name=banner.get("company_name"),
                    account_code=banner.get("account_code"),
                    account_name=banner.get("account_name"),
                    source_filename=excel_path.name,
                )
                customer_name = metadata.account_name or "Belirsiz Cari"

                for row_index in range(header_row + 1, sheet.max_row + 1):
                    description = clean_text(sheet.cell(row_index, colmap["AÇIKLAMA"]).value)
                    transaction_raw = sheet.cell(row_index, colmap["TARİH"]).value
                    if transaction_raw is None and description is None:
                        continue

                    voucher_no = clean_text(sheet.cell(row_index, colmap["FİŞ NO"]).value) if "FİŞ NO" in colmap else None
                    document_raw = sheet.cell(row_index, colmap["EVRAK TARİHİ"]).value if "EVRAK TARİHİ" in colmap else None
                    document_no = clean_text(sheet.cell(row_index, colmap["EVRAK NO"]).value) if "EVRAK NO" in colmap else None
                    debit = parse_decimal(sheet.cell(row_index, colmap["BORÇ"]).value) if "BORÇ" in colmap else Decimal("0")
                    credit = parse_decimal(sheet.cell(row_index, colmap["ALACAK"]).value) if "ALACAK" in colmap else Decimal("0")
                    balance = parse_decimal(sheet.cell(row_index, colmap["BAKİYE"]).value) if "BAKİYE" in colmap else Decimal("0")
                    balance_type = clean_text(sheet.cell(row_index, colmap["B/A"]).value) if "B/A" in colmap else None
                    movement_type = "Çek" if description and is_received_check_line(description) else "Ödeme / Hareket"

                    movements.append(
                        CustomerMovement(
                            customer_name=customer_name,
                            account_code=metadata.account_code,
                            source_file=excel_path.name,
                            source_sheet=sheet_name,
                            source_row_index=row_index,
                            transaction_date=parse_excel_date(transaction_raw),
                            voucher_no=voucher_no,
                            document_date=parse_excel_date(document_raw),
                            document_no=document_no,
                            movement_type=movement_type,
                            description=description,
                            debit=debit,
                            credit=credit,
                            balance=balance,
                            balance_type=balance_type,
                        )
                    )

                    if not description or not is_received_check_line(description):
                        continue

                    excel_amount = pick_movement_amount(debit, credit)
                    synthetic_line = build_synthetic_muavin_line(
                        transaction_raw,
                        voucher_no,
                        document_raw,
                        document_no,
                        description,
                        excel_amount,
                    )
                    record, _warnings, _debug = parse_received_check_line(synthetic_line, row_index, metadata)
                    if record is None:
                        continue
                    checks.append(
                        CustomerCheck(
                            customer_name=customer_name,
                            account_code=metadata.account_code,
                            source_file=excel_path.name,
                            source_row_index=row_index,
                            transaction_date=record.transaction_date,
                            check_no=record.check_no,
                            bank=record.bank,
                            maturity_date=record.maturity_date,
                            amount=record.amount or Decimal("0"),
                            description=record.raw_description,
                            warning=record.parse_warning,
                        )
                    )
        finally:
            workbook.close()

    movements.sort(key=lambda item: (item.customer_name, item.transaction_date or date.min, item.source_row_index))
    checks.sort(key=lambda item: (item.customer_name, item.maturity_date, item.check_no))
    return movements, checks, len(files)


def write_customer_payment_check_report(
    input_dir: str | Path,
    output_path: str | Path,
) -> CustomerReportResult:
    movements, checks, file_count = parse_customer_statement_excels(input_dir)
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_customer_summary_sheet(workbook, movements, checks)
    add_customer_movements_sheet(workbook, movements)
    add_customer_checks_sheet(workbook, checks)
    workbook.save(path)

    return CustomerReportResult(
        output_path=path,
        file_count=file_count,
        customer_count=len({movement.customer_name for movement in movements} | {check.customer_name for check in checks}),
        movement_count=len(movements),
        check_count=len(checks),
        check_total=sum((check.amount for check in checks), Decimal("0")),
    )


def add_customer_summary_sheet(workbook: Workbook, movements: list[CustomerMovement], checks: list[CustomerCheck]) -> None:
    sheet = workbook.create_sheet("Müşteri Özeti")
    append_header(
        sheet,
        [
            "Cari / Müşteri",
            "Hesap Kodu",
            "Hareket Adedi",
            "Borç Toplamı",
            "Alacak Toplamı",
            "Net Bakiye",
            "Çek Adedi",
            "Çek Toplamı",
        ],
    )

    grouped: dict[str, dict[str, Any]] = {}
    for movement in movements:
        item = grouped.setdefault(
            movement.customer_name,
            {
                "account_code": movement.account_code,
                "movement_count": 0,
                "debit": Decimal("0"),
                "credit": Decimal("0"),
                "check_count": 0,
                "check_total": Decimal("0"),
            },
        )
        item["movement_count"] += 1
        item["debit"] += movement.debit
        item["credit"] += movement.credit

    for check in checks:
        item = grouped.setdefault(
            check.customer_name,
            {
                "account_code": check.account_code,
                "movement_count": 0,
                "debit": Decimal("0"),
                "credit": Decimal("0"),
                "check_count": 0,
                "check_total": Decimal("0"),
            },
        )
        item["check_count"] += 1
        item["check_total"] += check.amount

    for customer_name, item in sorted(grouped.items()):
        net_balance = item["debit"] - item["credit"]
        sheet.append(
            [
                customer_name,
                item["account_code"],
                item["movement_count"],
                float(item["debit"]),
                float(item["credit"]),
                float(net_balance),
                item["check_count"],
                float(item["check_total"]),
            ]
        )

    sheet.append(
        [
            "GENEL TOPLAM",
            "",
            sum(item["movement_count"] for item in grouped.values()),
            float(sum((item["debit"] for item in grouped.values()), Decimal("0"))),
            float(sum((item["credit"] for item in grouped.values()), Decimal("0"))),
            float(sum((item["debit"] - item["credit"] for item in grouped.values()), Decimal("0"))),
            sum(item["check_count"] for item in grouped.values()),
            float(sum((item["check_total"] for item in grouped.values()), Decimal("0"))),
        ]
    )
    fill_row(sheet, sheet.max_row, TOTAL_FILL)
    format_sheet(sheet)
    format_columns(sheet, amount_columns=[4, 5, 6, 8])


def add_customer_movements_sheet(workbook: Workbook, movements: list[CustomerMovement]) -> None:
    sheet = workbook.create_sheet("Müşteri Bazlı Hareketler")
    append_header(
        sheet,
        [
            "Cari / Müşteri",
            "Hesap Kodu",
            "Tarih",
            "Fiş No",
            "Evrak Tarihi",
            "Evrak No",
            "Hareket Tipi",
            "Açıklama",
            "Borç",
            "Alacak",
            "Bakiye",
            "B/A",
            "Kaynak Dosya",
            "Satır",
        ],
    )
    for movement in movements:
        sheet.append(
            [
                movement.customer_name,
                movement.account_code,
                movement.transaction_date,
                movement.voucher_no,
                movement.document_date,
                movement.document_no,
                movement.movement_type,
                movement.description,
                float(movement.debit),
                float(movement.credit),
                float(movement.balance),
                movement.balance_type,
                movement.source_file,
                movement.source_row_index,
            ]
        )
    format_sheet(sheet)
    format_columns(sheet, date_columns=[3, 5], amount_columns=[9, 10, 11])


def add_customer_checks_sheet(workbook: Workbook, checks: list[CustomerCheck]) -> None:
    sheet = workbook.create_sheet("Müşteri Bazlı Çekler")
    append_header(
        sheet,
        [
            "Cari / Müşteri",
            "Hesap Kodu",
            "İşlem Tarihi",
            "Çek No",
            "Banka",
            "Vade Tarihi",
            "Tutar",
            "Açıklama",
            "Kaynak Dosya",
            "Satır",
            "Uyarı",
        ],
    )
    for check in checks:
        sheet.append(
            [
                check.customer_name,
                check.account_code,
                check.transaction_date,
                check.check_no,
                check.bank,
                check.maturity_date,
                float(check.amount),
                check.description,
                check.source_file,
                check.source_row_index,
                check.warning,
            ]
        )
    sheet.append(["GENEL TOPLAM", "", "", "", "", "", float(sum((check.amount for check in checks), Decimal("0"))), "", "", "", ""])
    fill_row(sheet, sheet.max_row, TOTAL_FILL)
    format_sheet(sheet)
    format_columns(sheet, date_columns=[3, 6], amount_columns=[7])


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def parse_excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_turkish_date(clean_text(value))


def parse_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def append_header(sheet: Any, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def fill_row(sheet: Any, row_number: int, fill: PatternFill) -> None:
    for cell in sheet[row_number]:
        cell.fill = fill
        cell.font = Font(bold=True)


def format_sheet(sheet: Any) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)
    if sheet.max_row > 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions


def format_columns(sheet: Any, date_columns: list[int] | None = None, amount_columns: list[int] | None = None) -> None:
    for column in date_columns or []:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column).number_format = DATE_FORMAT
    for column in amount_columns or []:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column).number_format = TRY_FORMAT


def format_customer_report_summary(result: CustomerReportResult) -> str:
    return (
        f"{result.file_count} Excel dosyası işlendi. "
        f"{result.customer_count} müşteri, {result.movement_count} hareket, "
        f"{result.check_count} çek bulundu. Çek toplamı: {format_try_amount(result.check_total)}"
    )
