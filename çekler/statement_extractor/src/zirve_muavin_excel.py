from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional

from openpyxl import Workbook

from .models import (
    CheckSourceType,
    ImportResult,
    ImportSession,
    ImportSessionStatus,
    ReviewStatus,
    StatementMetadata,
    StoredCheckRecord,
)
from .parser import is_received_check_line, parse_received_check_line


def extract_checks_from_zirve_muavin_workbook(
    workbook: Workbook,
    source_filename: str,
    import_session_id: str,
    source_type: CheckSourceType,
    review_status: ReviewStatus,
) -> ImportResult | None:
    """Zirve (ve benzeri) muavin Excel çıktısından ALINAN ÇEK satırlarını çıkarır.

    Tanınmayan çalışma kitapları için ``None`` döner (standart çek listesi yoluna düşülür).
    """
    combined_records: list[StoredCheckRecord] = []
    combined_warnings = []
    total_data_rows = 0
    statement_metadata: Optional[StatementMetadata] = None
    matched_any_sheet = False

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header = find_muavin_header_row(sheet)
        if not header:
            continue
        matched_any_sheet = True
        header_row, colmap = header
        banner = parse_zirve_banner(sheet, header_row)
        statement_metadata = StatementMetadata(
            company_name=banner.get("company_name"),
            account_code=banner.get("account_code"),
            account_name=banner.get("account_name"),
            source_filename=source_filename,
        )
        total_data_rows += count_muavin_data_rows(sheet, header_row, colmap)
        for row_index in range(header_row + 1, sheet.max_row + 1):
            aciklama_cell = sheet.cell(row_index, colmap["AÇIKLAMA"]).value
            if aciklama_cell is None or not str(aciklama_cell).strip():
                continue
            aciklama = str(aciklama_cell).strip()
            if not is_received_check_line(aciklama):
                continue
            tarih = sheet.cell(row_index, colmap["TARİH"]).value
            fis = sheet.cell(row_index, colmap["FİŞ NO"]).value if "FİŞ NO" in colmap else None
            evrak_tarih = sheet.cell(row_index, colmap["EVRAK TARİHİ"]).value if "EVRAK TARİHİ" in colmap else None
            evrak_no = sheet.cell(row_index, colmap["EVRAK NO"]).value if "EVRAK NO" in colmap else None
            borc = sheet.cell(row_index, colmap["BORÇ"]).value if "BORÇ" in colmap else None
            alacak = sheet.cell(row_index, colmap["ALACAK"]).value if "ALACAK" in colmap else None
            excel_amount = pick_movement_amount(borc, alacak)
            synthetic = build_synthetic_muavin_line(tarih, fis, evrak_tarih, evrak_no, aciklama, excel_amount)
            record, line_warnings, _debug = parse_received_check_line(synthetic, row_index, statement_metadata)
            combined_warnings.extend(line_warnings)
            if record is None:
                continue
            if record.amount is None and excel_amount is not None:
                record = record.model_copy(update={"amount": excel_amount})
            combined_records.append(
                StoredCheckRecord(
                    **record.model_dump(),
                    source_type=source_type,
                    source_file=source_filename,
                    import_session_id=import_session_id,
                    review_status=review_status,
                    source_row_index=row_index,
                    source_sheet=sheet_name,
                )
            )

    if not matched_any_sheet:
        return None

    meta = statement_metadata or StatementMetadata(source_filename=source_filename)
    session = ImportSession(
        import_session_id=import_session_id,
        source_type=source_type,
        source_file=source_filename,
        status=ImportSessionStatus.PARSED,
        statement_metadata=meta,
        total_rows=total_data_rows,
        parsed_count=len(combined_records),
        warning_count=len(combined_warnings),
        error_count=0,
        completed_at=datetime.utcnow(),
    )
    return ImportResult(import_session=session, records=combined_records, warnings=combined_warnings, errors=[])


def find_muavin_header_row(sheet: Any) -> Optional[tuple[int, dict[str, int]]]:
    max_scan = min(sheet.max_row, 35)
    for row_index in range(1, max_scan + 1):
        colmap: dict[str, int] = {}
        for column_index in range(1, sheet.max_column + 1):
            value = sheet.cell(row_index, column_index).value
            if value is None or str(value).strip() == "":
                continue
            key = str(value).strip()
            colmap[key] = column_index
        if "TARİH" in colmap and "AÇIKLAMA" in colmap and "FİŞ NO" in colmap:
            return row_index, colmap
    return None


def count_muavin_data_rows(sheet: Any, header_row: int, colmap: dict[str, int]) -> int:
    count = 0
    for row_index in range(header_row + 1, sheet.max_row + 1):
        tarih = sheet.cell(row_index, colmap["TARİH"]).value
        aciklama = sheet.cell(row_index, colmap["AÇIKLAMA"]).value
        if tarih is not None or (aciklama is not None and str(aciklama).strip()):
            count += 1
    return count


def parse_zirve_banner(sheet: Any, header_row: int) -> dict[str, Optional[str]]:
    company_name: Optional[str] = None
    account_code: Optional[str] = None
    account_name: Optional[str] = None
    for row_index in range(1, header_row):
        for column_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_index, column_index).value
            if cell is None:
                continue
            label = str(cell).strip()
            if not label:
                continue
            upper = label.upper()
            if upper.startswith("FIRMA") and ":" in label:
                company_name = label.split(":", 1)[1].strip()
            if "ALT HESAP KODU" in upper and ":" in label:
                code = label.split(":", 1)[1].strip()
                if code:
                    account_code = code
                else:
                    neighbor = sheet.cell(row_index, column_index + 1).value
                    if neighbor is not None and str(neighbor).strip():
                        account_code = str(neighbor).strip()
            if "ALT HESAP AD" in upper and ":" in label:
                name = label.split(":", 1)[1].strip()
                if name:
                    account_name = name
                else:
                    for offset in (1, 2, 3):
                        neighbor = sheet.cell(row_index, column_index + offset).value
                        if neighbor is not None and str(neighbor).strip():
                            account_name = str(neighbor).strip()
                            break
    return {
        "company_name": company_name,
        "account_code": account_code,
        "account_name": account_name,
    }


def pick_movement_amount(borc: Any, alacak: Any) -> Optional[Decimal]:
    b = _excel_decimal(borc)
    a = _excel_decimal(alacak)
    if a is not None and a != 0:
        return abs(a)
    if b is not None and b != 0:
        return abs(b)
    return None


def _excel_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    return None


def excel_cell_to_date_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def excel_cell_to_token(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal) and value == value.to_integral_value():
        return str(int(value))
    return str(value).strip()


def format_turkish_amount_for_line(amount: Decimal) -> str:
    quantized = amount.quantize(Decimal("0.01"))
    negative = quantized < 0
    quantized = abs(quantized)
    whole = int(quantized)
    frac = int((quantized - whole) * 100)
    if frac >= 100:
        whole += 1
        frac -= 100

    def group_int(n: int) -> str:
        s = str(abs(n))
        parts: list[str] = []
        while s:
            parts.append(s[-3:])
            s = s[:-3]
        return ".".join(reversed(parts))

    body = group_int(whole)
    out = f"{body},{frac:02d}"
    return f"-{out}" if negative else out


def build_synthetic_muavin_line(
    tarih: Any,
    fis_no: Any,
    evrak_tarih: Any,
    evrak_no: Any,
    aciklama: str,
    amount: Optional[Decimal],
) -> str:
    parts = [
        excel_cell_to_date_str(tarih),
        excel_cell_to_token(fis_no),
        excel_cell_to_date_str(evrak_tarih),
        excel_cell_to_token(evrak_no),
        aciklama,
    ]
    line = " ".join(p for p in parts if p).strip()
    if amount is not None:
        line = f"{line} {format_turkish_amount_for_line(amount)}"
    return line
