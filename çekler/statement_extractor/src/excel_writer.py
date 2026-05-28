from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import CheckRecord, ParseResult
from .utils import calculate_days_until_maturity, calculate_maturity_status, get_week_period


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9EAD3")
OVERDUE_FILL = PatternFill("solid", fgColor="F4CCCC")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
TRY_FORMAT = '#,##0.00 "TL"'
DATE_FORMAT = "dd.mm.yyyy"


def write_parse_result_to_excel(
    parse_result: ParseResult,
    output_path: str | Path,
    today: date | None = None,
    include_debug: bool = True,
    include_overdue_in_value_calc: bool = False,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_check_list_sheet(workbook, parse_result, today=today)
    add_summary_sheet(workbook, parse_result)
    add_bank_summary_sheet(workbook, parse_result)
    add_maturity_calendar_sheet(workbook, parse_result)
    add_value_calculation_sheet(
        workbook,
        parse_result,
        today=today,
        include_overdue=include_overdue_in_value_calc,
    )
    if include_debug:
        add_debug_sheet(workbook, parse_result)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def add_check_list_sheet(workbook: Workbook, parse_result: ParseResult, today: date | None = None) -> None:
    sheet = workbook.create_sheet("Çek Listesi")
    headers = [
        "Sıra",
        "İşlem Tarihi",
        "Evrak Tarihi",
        "Fiş No",
        "Evrak No",
        "Hareket Tipi",
        "Çek No",
        "Banka",
        "Vade Tarihi",
        "Vade Ayı",
        "Vadeye Kalan Gün",
        "Vade Durumu",
        "Haftalık Vade Periyodu",
        "Tutar",
        "Para Birimi",
        "Gönderildiği Yer",
        "Açıklama",
        "Kaynak Sayfa",
        "Parse Uyarısı",
    ]
    append_header(sheet, headers)

    total = Decimal("0")
    for index, check in enumerate(parse_result.checks, start=1):
        total += check.amount or Decimal("0")
        days = calculate_days_until_maturity(check.maturity_date, today)
        row = [
            index,
            check.transaction_date,
            check.document_date,
            check.voucher_no,
            check.document_no,
            check.movement_type.value,
            check.check_no,
            check.bank,
            check.maturity_date,
            check.maturity_month,
            days,
            calculate_maturity_status(check.maturity_date, today),
            get_week_period(check.maturity_date),
            float(check.amount or Decimal("0")),
            check.currency,
            check.sent_to,
            check.raw_description,
            check.source_page,
            check.parse_warning,
        ]
        sheet.append(row)
        row_number = sheet.max_row
        if days < 0:
            fill_row(sheet, row_number, OVERDUE_FILL)
        if check.parse_warning:
            fill_row(sheet, row_number, WARNING_FILL)

    sheet.append(["Toplam"] + [""] * 12 + [float(total)] + [""] * 5)
    fill_row(sheet, sheet.max_row, TOTAL_FILL)
    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)

    format_common_sheet(sheet)
    for column in (2, 3, 9):
        format_column(sheet, column, DATE_FORMAT)
    format_column(sheet, 14, TRY_FORMAT)
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"


def add_summary_sheet(workbook: Workbook, parse_result: ParseResult) -> None:
    sheet = workbook.create_sheet("Özet")
    metadata = parse_result.statement_metadata
    summary = parse_result.summary
    rows = [
        ("Firma adı", metadata.company_name),
        ("Alt hesap kodu", metadata.account_code),
        ("Alt hesap adı", metadata.account_name),
        ("PDF sayfa sayısı", metadata.page_count),
        ("Toplam çek adedi", summary.total_count),
        ("Toplam çek tutarı", float(summary.total_amount)),
        ("En erken vade", summary.earliest_maturity_date),
        ("En geç vade", summary.latest_maturity_date),
        ("Vadesi geçmiş çek toplamı", float(summary.overdue_total)),
        ("7 gün içinde vadesi gelecek çek toplamı", float(summary.due_in_7_days_total)),
        ("30 gün içinde vadesi gelecek çek toplamı", float(summary.due_in_30_days_total)),
        ("Parse uyarısı sayısı", summary.warning_count),
    ]
    for row in rows:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 42
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=1).font = Font(bold=True)
    for row in (6, 9, 10, 11):
        sheet.cell(row=row, column=2).number_format = TRY_FORMAT
    for row in (7, 8):
        sheet.cell(row=row, column=2).number_format = DATE_FORMAT


def add_bank_summary_sheet(workbook: Workbook, parse_result: ParseResult) -> None:
    sheet = workbook.create_sheet("Banka Özeti")
    append_header(sheet, ["Banka", "Çek adedi", "Toplam tutar", "Ortalama tutar", "En erken vade", "En geç vade"])
    grouped: dict[str, list[CheckRecord]] = defaultdict(list)
    for check in parse_result.checks:
        grouped[check.bank].append(check)

    for bank, checks in sorted(grouped.items()):
        total = sum((check.amount or Decimal("0") for check in checks), Decimal("0"))
        maturities = [check.maturity_date for check in checks]
        sheet.append([bank, len(checks), float(total), float(total / len(checks)), min(maturities), max(maturities)])

    format_common_sheet(sheet)
    format_column(sheet, 3, TRY_FORMAT)
    format_column(sheet, 4, TRY_FORMAT)
    format_column(sheet, 5, DATE_FORMAT)
    format_column(sheet, 6, DATE_FORMAT)


def add_maturity_calendar_sheet(workbook: Workbook, parse_result: ParseResult) -> None:
    sheet = workbook.create_sheet("Vade Takvimi")
    append_header(sheet, ["Vade Ayı", "Adet", "Toplam"])
    month_counts: dict[str, int] = defaultdict(int)
    week_counts: dict[str, int] = defaultdict(int)
    for check in parse_result.checks:
        month_counts[check.maturity_month] += 1
        week_counts[get_week_period(check.maturity_date)] += 1

    for month, total in sorted(parse_result.summary.maturity_month_totals.items()):
        sheet.append([month, month_counts[month], float(total)])

    sheet.append([])
    append_header(sheet, ["Haftalık Vade Periyodu", "Adet", "Toplam"])
    for week, total in sorted(parse_result.summary.weekly_maturity_totals.items()):
        sheet.append([week, week_counts[week], float(total)])

    format_common_sheet(sheet)
    format_column(sheet, 3, TRY_FORMAT)


def add_value_calculation_sheet(
    workbook: Workbook,
    parse_result: ParseResult,
    today: date | None = None,
    include_overdue: bool = False,
) -> None:
    sheet = workbook.create_sheet("Değer Hesabı")
    append_header(sheet, ["Sıra", "Çek No", "Banka", "Vade Tarihi", "Vadeye Kalan Gün", "Tutar", "Gün × Tutar"])

    total_amount = Decimal("0")
    total_weighted_days = Decimal("0")
    excluded_overdue_count = 0

    for index, check in enumerate(parse_result.checks, start=1):
        amount = check.amount or Decimal("0")
        days = calculate_days_until_maturity(check.maturity_date, today)
        if days < 0 and not include_overdue:
            excluded_overdue_count += 1
            continue

        weighted_days = Decimal(days) * amount
        total_amount += amount
        total_weighted_days += weighted_days
        sheet.append(
            [
                index,
                check.check_no,
                check.bank,
                check.maturity_date,
                days,
                float(amount),
                float(weighted_days),
            ]
        )

    average_maturity = total_weighted_days / total_amount if total_amount else Decimal("0")
    sheet.append([])
    total_start = sheet.max_row + 1
    sheet.append(["Toplam Tutar", "", "", "", "", float(total_amount), ""])
    sheet.append(["Toplam Gün × Tutar", "", "", "", "", "", float(total_weighted_days)])
    sheet.append(["Ağırlıklı Ortalama Vade", "", "", "", f"{average_maturity:.2f} gün", "", ""])
    if not include_overdue:
        sheet.append(
            [
                "Not",
                "Vadesi geçmiş çekler ağırlıklı ortalama vade hesabına dahil edilmemiştir.",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        if excluded_overdue_count:
            sheet.append(["Hariç tutulan vadesi geçmiş çek adedi", excluded_overdue_count, "", "", "", "", ""])

    for row_number in range(total_start, sheet.max_row + 1):
        fill_row(sheet, row_number, TOTAL_FILL)
        sheet.cell(row=row_number, column=1).font = Font(bold=True)

    format_common_sheet(sheet)
    format_column(sheet, 4, DATE_FORMAT)
    format_column(sheet, 6, TRY_FORMAT)
    format_column(sheet, 7, '#,##0.00')


def calculate_weighted_average_maturity(
    checks: list[CheckRecord],
    today: date | None = None,
    include_overdue: bool = False,
) -> Decimal:
    total_amount = Decimal("0")
    total_weighted_days = Decimal("0")
    for check in checks:
        amount = check.amount or Decimal("0")
        days = calculate_days_until_maturity(check.maturity_date, today)
        if days < 0 and not include_overdue:
            continue
        total_amount += amount
        total_weighted_days += Decimal(days) * amount
    return total_weighted_days / total_amount if total_amount else Decimal("0")


def add_debug_sheet(workbook: Workbook, parse_result: ParseResult) -> None:
    sheet = workbook.create_sheet("Parse Debug")
    append_header(
        sheet,
        [
            "Kaynak sayfa",
            "Ham satır",
            "Parse edildi mi?",
            "Bulunan çek no",
            "Bulunan banka",
            "Bulunan vade",
            "Bulunan para değerleri",
            "Kullanılan tutar",
            "Uyarı",
        ],
    )
    for row in parse_result.debug_rows:
        sheet.append(
            [
                row.source_page,
                row.raw_line,
                "Evet" if row.parsed else "Hayır",
                row.found_check_no,
                row.found_bank,
                row.found_maturity_date,
                ", ".join(str(amount) for amount in row.found_amounts),
                float(row.used_amount) if row.used_amount is not None else None,
                row.warning,
            ]
        )
    format_common_sheet(sheet)
    format_column(sheet, 6, DATE_FORMAT)
    format_column(sheet, 8, TRY_FORMAT)


def append_header(sheet: Any, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def fill_row(sheet: Any, row_number: int, fill: PatternFill) -> None:
    for cell in sheet[row_number]:
        cell.fill = fill


def format_common_sheet(sheet: Any) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    auto_width(sheet)
    if sheet.max_row > 1:
        sheet.freeze_panes = "A2"


def format_column(sheet: Any, column_index: int, number_format: str) -> None:
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=column_index).number_format = number_format


def auto_width(sheet: Any) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
