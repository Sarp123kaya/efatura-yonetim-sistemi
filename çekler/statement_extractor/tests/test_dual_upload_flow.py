from __future__ import annotations

from datetime import date
from decimal import Decimal

from openpyxl import Workbook, load_workbook

from statement_extractor.src.consolidated_report import write_consolidated_check_report
from statement_extractor.src.customer_report import parse_customer_statement_excels, write_customer_payment_check_report
from statement_extractor.src.excel_importer import extract_checks_from_excel
from statement_extractor.src.image_ocr_reader import parse_check_from_ocr_text
from statement_extractor.src.models import CheckSourceType, ImportResult, ImportSession, ReviewStatus, StoredCheckRecord
from statement_extractor.src.postgres_repository import PostgresCheckRepository, row_to_stored_check


def write_upload_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Çek Listesi"
    sheet.append(["Sıra", "Çek No", "Banka", "Vade Tarihi", "Tutar", "Lehtar / Alacaklı", "Keşideci / Firma", "Gönderildiği Yer", "Açıklama"])
    sheet.append([1, "123456", "HALKBANK", "25.05.2026", 2550000, "Cari A", "Firma A", "Banka Teminat", "Excel satırı"])
    workbook.save(path)


def test_excel_importer_extracts_stored_check_record(tmp_path) -> None:
    excel_path = tmp_path / "checks.xlsx"
    write_upload_workbook(excel_path)

    result = extract_checks_from_excel(excel_path)

    assert result.import_session.source_type == CheckSourceType.EXCEL
    assert len(result.records) == 1
    record = result.records[0]
    assert record.check_no == "123456"
    assert record.bank == "Halkbank"
    assert record.maturity_date == date(2026, 5, 25)
    assert record.amount == Decimal("2550000")
    assert record.sent_to == "Banka Teminat"
    assert record.source_type == CheckSourceType.EXCEL


def write_zirve_muavin_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Zirve"
    sheet.cell(2, 1, "Firma : Test Limited")
    sheet.cell(3, 1, "Alt Hesap Kodu :")
    sheet.cell(3, 2, "120.01.001")
    sheet.cell(3, 3, "Alt Hesap Adı :")
    sheet.cell(3, 4, "Örnek Cari")
    headers = [
        "TARİH",
        "FİŞ NO",
        "MADDE NO",
        "EVRAK TARİHİ",
        "SERİ NO",
        "EVRAK NO",
        "AÇIKLAMA",
        "BORÇ",
        "ALACAK",
        "BAKİYE",
        "B/A",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(4, column_index, header)
    sheet.cell(5, 1, "20.01.2026")
    sheet.cell(5, 2, "000000011")
    sheet.cell(5, 4, "20.01.2026")
    sheet.cell(5, 6, "2001")
    sheet.cell(5, 7, "ALINAN ÇEK (5526515 NL HALKBANK VD:25.05.26)")
    sheet.cell(5, 8, 0)
    sheet.cell(5, 9, 2550000)
    workbook.save(path)


def test_zirve_muavin_excel_extracts_received_checks(tmp_path) -> None:
    excel_path = tmp_path / "zirve.xlsx"
    write_zirve_muavin_workbook(excel_path)

    result = extract_checks_from_excel(excel_path)

    assert len(result.records) == 1
    record = result.records[0]
    assert record.check_no == "5526515"
    assert record.bank == "Halkbank"
    assert record.maturity_date == date(2026, 5, 25)
    assert record.amount == Decimal("2550000")
    assert record.source_row_index == 5
    assert record.source_sheet == "Zirve"


def test_customer_report_parses_statements_and_writes_excel(tmp_path) -> None:
    input_dir = tmp_path / "uploads"
    input_dir.mkdir()
    excel_path = input_dir / "zirve.xlsx"
    output_path = tmp_path / "customer_report.xlsx"
    write_zirve_muavin_workbook(excel_path)

    movements, checks, file_count = parse_customer_statement_excels(input_dir)
    result = write_customer_payment_check_report(input_dir, output_path)

    assert file_count == 1
    assert len(movements) == 1
    assert len(checks) == 1
    assert result.movement_count == 1
    assert result.check_count == 1
    assert result.customer_count == 1

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "Müşteri Özeti",
        "Müşteri Bazlı Hareketler",
        "Müşteri Bazlı Çekler",
    ]
    assert workbook["Müşteri Özeti"].cell(row=2, column=1).value == "Örnek Cari"
    assert workbook["Müşteri Bazlı Hareketler"].cell(row=2, column=7).value == "Çek"
    assert workbook["Müşteri Bazlı Çekler"].cell(row=2, column=4).value == "5526515"


def test_ocr_text_parser_extracts_draft_record() -> None:
    text = """
    Garanti BBVA
    Çek No: 801373
    Keşide Tarihi 20.08.2026
    #1.500.000#
    """

    record, warning = parse_check_from_ocr_text(text, "image.jpeg", "session-1")

    assert warning is None
    assert record is not None
    assert record.check_no == "801373"
    assert record.bank == "Garanti BBVA"
    assert record.maturity_date == date(2026, 8, 20)
    assert record.amount == Decimal("1500000.00")
    assert record.review_status == ReviewStatus.NEEDS_REVIEW


def test_ocr_text_parser_warns_when_required_fields_missing() -> None:
    record, warning = parse_check_from_ocr_text("Garanti BBVA sadece banka var", "image.jpeg", "session-1")

    assert record is None
    assert warning is not None
    assert warning.warning_type == "ocr_low_confidence"


def test_consolidated_report_writes_expected_sheets(tmp_path) -> None:
    output_path = tmp_path / "all_checks.xlsx"
    records = [
        StoredCheckRecord(
            check_no="123456",
            bank="Halkbank",
            maturity_date=date(2026, 4, 25),
            maturity_month="2026-04",
            amount=Decimal("2550000"),
            customer_name="Cari A",
            company_name="Firma A",
            sent_to="Banka Teminat",
            source_page=1,
            raw_description="Test",
            raw_line="Test",
            source_type=CheckSourceType.EXCEL,
            source_file="checks.xlsx",
            review_status=ReviewStatus.APPROVED,
        ),
        StoredCheckRecord(
            check_no="801373",
            bank="Garanti BBVA",
            maturity_date=date(2026, 8, 20),
            maturity_month="2026-08",
            amount=Decimal("1500000"),
            customer_name="Cari B",
            company_name="Firma B",
            sent_to="Tedarikçi B",
            source_page=1,
            raw_description="OCR",
            raw_line="OCR",
            source_type=CheckSourceType.IMAGE_OCR,
            source_file="image.jpeg",
            review_status=ReviewStatus.NEEDS_REVIEW,
        ),
    ]

    write_consolidated_check_report(records, output_path, today=date(2026, 5, 3))
    workbook = load_workbook(output_path)

    assert workbook.sheetnames == [
        "Tüm Çekler",
        "Cari Bazlı Özet",
        "Banka Özeti",
        "Vade Takvimi",
        "Değer Hesabı",
        "Kaynak ve Kontrol Geçmişi",
    ]
    assert workbook["Tüm Çekler"].max_row == 4
    all_checks = workbook["Tüm Çekler"]
    assert all_checks.cell(row=1, column=12).value == "Gönderildiği Yer"
    assert all_checks.cell(row=2, column=12).value == "Banka Teminat"
    assert all_checks.cell(row=3, column=12).value == "Tedarikçi B"
    account_summary = workbook["Cari Bazlı Özet"]
    assert [account_summary.cell(row=1, column=column).value for column in range(1, 8)] == [
        "Cari / Lehtar",
        "Çek adedi",
        "Toplam tutar",
        "Ödenen çek adedi",
        "Ödenen tutar",
        "En erken vade",
        "En geç vade",
    ]
    assert account_summary.cell(row=2, column=4).value == 1
    assert account_summary.cell(row=2, column=5).value == 2550000


def test_repository_import_result_keeps_duplicate_session_rows() -> None:
    record = StoredCheckRecord(
        check_no="123456",
        bank="Halkbank",
        maturity_date=date(2026, 5, 25),
        maturity_month="2026-05",
        amount=Decimal("2550000"),
        customer_name="Cari A",
        company_name="Firma A",
        source_page=1,
        raw_description="Test",
        raw_line="Test",
        source_type=CheckSourceType.EXCEL,
        source_file="checks.xlsx",
        import_session_id="session-2",
        review_status=ReviewStatus.APPROVED,
    )
    result = ImportResult(
        import_session=ImportSession(
            import_session_id="session-2",
            source_type=CheckSourceType.EXCEL,
            source_file="checks.xlsx",
        ),
        records=[record],
    )
    repository = FakeRepository(existing_check_id=42)

    summary = repository.import_result(result)

    assert summary["inserted_count"] == 0
    assert summary["skipped_duplicate_count"] == 1
    assert summary["import_row_count"] == 1
    assert repository.import_rows == [{"canonical": None, "duplicate": 42}]


def test_row_to_stored_check_maps_sent_to() -> None:
    record = row_to_stored_check(
        (
            "RECEIVED_CHECK",
            "123456",
            "Halkbank",
            Decimal("2550000"),
            "TRY",
            date(2026, 5, 25),
            None,
            None,
            None,
            None,
            "Cari A",
            "Firma A",
            "checks.xlsx",
            "Banka Teminat",
            1,
            "Test",
            "Test",
            None,
            "EXCEL",
            "11111111-1111-1111-1111-111111111111",
            "APPROVED",
            2,
            "Çek Listesi",
            None,
            None,
        )
    )

    assert record.sent_to == "Banka Teminat"
    assert record.source_page == 1
    assert record.source_type == CheckSourceType.EXCEL


class FakeRepository(PostgresCheckRepository):
    def __init__(self, existing_check_id):
        self.existing_check_id = existing_check_id
        self.import_rows = []
        self.sessions = []

    def upsert_import_session(self, session):
        self.sessions.append(session)

    def find_existing_check_id(self, record):
        return self.existing_check_id

    def insert(self, record, source_file=None):
        return 100

    def insert_import_row(self, record, canonical_check_id=None, duplicate_of_check_id=None):
        self.import_rows.append({"canonical": canonical_check_id, "duplicate": duplicate_of_check_id})
