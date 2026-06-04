#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export Database to Excel
Export incoming and outgoing invoices to Excel files
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from datetime import datetime
import json
from openpyxl import load_workbook

from backend.core.db import db
from report_cleanup import cleanup_old_reports

def format_jsonb_for_excel(data):
    """Format JSONB data for Excel"""
    if not data:
        return ""
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except:
            return data
    if isinstance(data, list):
        return ", ".join(str(x) for x in data)
    return str(data)


def save_excel_with_number_formats(df, output_file, number_columns):
    """Save DataFrame and keep selected columns as numeric Excel cells."""
    df.to_excel(output_file, index=False, engine='openpyxl')
    wb = load_workbook(output_file)
    ws = wb.active
    header_to_idx = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value
    }
    for column in number_columns:
        col_idx = header_to_idx.get(column)
        if not col_idx:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
    wb.save(output_file)

def export_incoming_invoices(output_file):
    """Export incoming invoices to Excel"""
    print("📥 Exporting incoming invoices...")
    cleanup_old_reports(Path(output_file).parent, ["Gelen_Faturalar_*.xlsx"])
    
    query = """
        SELECT 
            invoice_id,
            uuid,
            issue_date,
            supplier,
            supplier_tckn_vkn,
            amount,
            total_vat_base,
            currency,
            despatch_ids,
            change_type,
            last_change_at,
            created_at,
            updated_at
        FROM incoming_invoices
        ORDER BY issue_date DESC
    """
    
    rows = db.query(query)
    
    if not rows:
        print("❌ No incoming invoices found")
        return None
    
    # Convert to DataFrame
    df = pd.DataFrame(rows)
    
    # Format JSONB columns
    df['despatch_ids'] = df['despatch_ids'].apply(format_jsonb_for_excel)
    
    # Format dates
    for col in ['issue_date', 'last_change_at', 'created_at', 'updated_at']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col])
    
    # Rename columns to Turkish
    df = df.rename(columns={
        'invoice_id': 'Fatura ID',
        'uuid': 'UUID',
        'issue_date': 'Düzenleme Tarihi',
        'supplier': 'Tedarikçi',
        'supplier_tckn_vkn': 'TCKN/VKN',
        'amount': 'Tutar',
        'total_vat_base': 'KDV Matrahı',
        'currency': 'Para Birimi',
        'despatch_ids': 'İrsaliye Kodları',
        'change_type': 'Değişiklik Tipi',
        'last_change_at': 'Son Değişiklik',
        'created_at': 'Oluşturma Tarihi',
        'updated_at': 'Güncelleme Tarihi'
    })
    df = df.drop(
        columns=[
            'UUID',
            'TCKN/VKN',
            'Para Birimi',
            'Değişiklik Tipi',
            'Son Değişiklik',
            'Oluşturma Tarihi',
        ],
        errors='ignore',
    )
    
    # Save to Excel
    save_excel_with_number_formats(df, output_file, ['Tutar', 'KDV Matrahı'])
    print(f"✅ Exported {len(df)} incoming invoices to {output_file}")
    
    return len(df)

def export_outgoing_invoices(output_file):
    """Export outgoing invoices to Excel"""
    print("📤 Exporting outgoing invoices...")
    cleanup_old_reports(Path(output_file).parent, ["Giden_Faturalar_*.xlsx"])
    
    query = """
        SELECT 
            id,
            invoice_no,
            issue_date,
            firm_name,
            total_tl,
            taxable_amount,
            description,
            COALESCE(irsaliye_codes_override, irsaliye_codes) AS irsaliye_codes,
            change_type,
            last_change_at,
            created_at,
            updated_at
        FROM outgoing_invoices
        ORDER BY issue_date DESC
    """
    
    rows = db.query(query)
    
    if not rows:
        print("❌ No outgoing invoices found")
        return None
    
    # Convert to DataFrame
    df = pd.DataFrame(rows)
    
    # Format JSONB columns
    df['irsaliye_codes'] = df['irsaliye_codes'].apply(format_jsonb_for_excel)
    
    # Format dates
    for col in ['issue_date', 'last_change_at', 'created_at', 'updated_at']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col])
    
    # Rename columns to Turkish
    df = df.rename(columns={
        'id': 'ID',
        'invoice_no': 'Fatura No',
        'issue_date': 'Düzenleme Tarihi',
        'firm_name': 'Firma Adı',
        'total_tl': 'Toplam (TL)',
        'taxable_amount': 'Vergiye Esas',
        'description': 'Açıklama',
        'irsaliye_codes': 'İrsaliye Kodları',
        'change_type': 'Değişiklik Tipi',
        'last_change_at': 'Son Değişiklik',
        'created_at': 'Oluşturma Tarihi',
        'updated_at': 'Güncelleme Tarihi'
    })
    df = df.drop(
        columns=[
            'Değişiklik Tipi',
            'Son Değişiklik',
            'Oluşturma Tarihi',
        ],
        errors='ignore',
    )
    
    # Save to Excel
    save_excel_with_number_formats(df, output_file, ['Toplam (TL)', 'Vergiye Esas'])
    print(f"✅ Exported {len(df)} outgoing invoices to {output_file}")
    
    return len(df)

def export_agent_runs(output_file):
    """Export agent runs to Excel"""
    print("🤖 Exporting agent runs...")
    cleanup_old_reports(Path(output_file).parent, ["Agent_Calismalari_*.xlsx"])
    
    query = """
        SELECT 
            id,
            agent_name,
            run_id,
            start_time,
            end_time,
            duration_sec,
            insert_count,
            update_count,
            nochange_count,
            error_count,
            total_fetched,
            batch_count,
            status,
            error_message,
            host,
            agent_version,
            created_at
        FROM agent_runs
        ORDER BY start_time DESC
    """
    
    rows = db.query(query)
    
    if not rows:
        print("❌ No agent runs found")
        return None
    
    # Convert to DataFrame
    df = pd.DataFrame(rows)
    
    # Format dates (remove timezone for Excel compatibility)
    for col in ['start_time', 'end_time', 'created_at']:
        if col in df.columns and df[col].dtype != 'object':
            try:
                if hasattr(df[col].dt, 'tz') and df[col].dt.tz is not None:
                    df[col] = df[col].dt.tz_localize(None)
                else:
                    df[col] = pd.to_datetime(df[col], utc=True).dt.tz_localize(None)
            except:
                df[col] = pd.to_datetime(df[col], errors='coerce')
    
    # Rename columns to Turkish
    df = df.rename(columns={
        'id': 'ID',
        'agent_name': 'Agent Adı',
        'run_id': 'Çalışma ID',
        'start_time': 'Başlangıç',
        'end_time': 'Bitiş',
        'duration_sec': 'Süre (sn)',
        'insert_count': 'Eklenen',
        'update_count': 'Güncellenen',
        'nochange_count': 'Değişmeyen',
        'error_count': 'Hata',
        'total_fetched': 'Çekilen Toplam',
        'batch_count': 'Batch Sayısı',
        'status': 'Durum',
        'error_message': 'Hata Mesajı',
        'host': 'Sunucu',
        'agent_version': 'Versiyon',
        'created_at': 'Oluşturma Tarihi'
    })
    
    # Save to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')
    print(f"✅ Exported {len(df)} agent runs to {output_file}")
    
    return len(df)

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Export database to Excel')
    parser.add_argument('--output-dir', default='kayıtlar', help='Output directory')
    parser.add_argument('--type', choices=['incoming', 'outgoing', 'runs', 'all'], 
                       default='all', help='Type of data to export')
    args = parser.parse_args()
    
    # Create output directory
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    print("\n" + "=" * 80)
    print("📊 EXCEL EXPORT")
    print("=" * 80)
    print()
    
    try:
        total = 0
        
        if args.type in ['all', 'incoming']:
            file = output_dir / f'Gelen_Faturalar_{timestamp}.xlsx'
            count = export_incoming_invoices(file)
            if count:
                total += count
        
        if args.type in ['all', 'outgoing']:
            file = output_dir / f'Giden_Faturalar_{timestamp}.xlsx'
            count = export_outgoing_invoices(file)
            if count:
                total += count
        
        if args.type in ['all', 'runs']:
            file = output_dir / f'Agent_Calismalari_{timestamp}.xlsx'
            count = export_agent_runs(file)
            if count:
                total += count
        
        print()
        print("=" * 80)
        print(f"✅ Export complete! Total records: {total}")
        print("=" * 80)
        print()
        print(f"📁 Files saved to: {output_dir.absolute()}")
        print()
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
