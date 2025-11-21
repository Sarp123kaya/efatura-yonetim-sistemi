#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Veritabanından Excel'e Export Script
"""

import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def create_excel_export():
    """Veritabanındaki verileri Excel dosyasına export eder"""
    
    # Proje kök dizinine git
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    os.chdir(project_root)
    
    # kayıtlar klasörünü oluştur
    excel_dir = 'data/excel/birlesik'
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
        print(f"✓ '{excel_dir}' klasörü oluşturuldu")
    
    # Veritabanına bağlan
    db_path = 'data/db/birlesik.db'
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Excel dosyası oluştur
    wb = openpyxl.Workbook()
    
    # Stil tanımlamaları
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== FATURALAR SAYFAsl ==========
    ws_invoices = wb.active
    ws_invoices.title = "Faturalar"
    
    # Başlıklar
    headers_invoices = [
        'Firma', 'Fatura No', 'Tarih', 'Toplam Tutar (TL)', 
        'Vergi Matrahı', 'KDV Tutarı', 'Satıcı Firma', 'Müşteri Firma', 'Açıklama'
    ]
    
    for col_num, header in enumerate(headers_invoices, 1):
        cell = ws_invoices.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Verileri ekle (DISTINCT ile tekrarları engelle)
    cursor.execute('''
        SELECT DISTINCT
            firma_kodu, invoice_number, issue_date, total_amount,
            taxable_amount, tax_amount, supplier_name, customer_name, description
        FROM invoices
        GROUP BY invoice_number
        ORDER BY firma_kodu, issue_date DESC
    ''')
    
    for row_num, row_data in enumerate(cursor.fetchall(), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_invoices.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Sayısal sütunları formatla
            if col_num in [4, 5, 6]:  # Tutar sütunları
                cell.number_format = '#,##0.00'
    
    # Sütun genişliklerini ayarla
    column_widths = [10, 20, 12, 18, 18, 18, 40, 40, 50]  # Description için 50 genişlik
    for i, width in enumerate(column_widths, 1):
        ws_invoices.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # ========== FATURA SATIRLARI SAYFAsl ==========
    ws_lines = wb.create_sheet("Fatura Satırları")
    
    headers_lines = [
        'Firma', 'Fatura No', 'Satır No', 'Ürün/Hizmet Adı', 
        'Miktar', 'Birim', 'Birim Fiyat', 'Satır Toplamı', 'ADET'
    ]
    
    for col_num, header in enumerate(headers_lines, 1):
        cell = ws_lines.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    cursor.execute('''
        SELECT 
            i.firma_kodu, i.invoice_number, il.line_id, il.item_name,
            il.quantity, il.unit, il.unit_price, il.line_total
        FROM invoice_lines il
        JOIN invoices i ON il.invoice_id = i.id
        ORDER BY i.firma_kodu, i.issue_date DESC, il.line_id
    ''')
    
    for row_num, row_data in enumerate(cursor.fetchall(), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_lines.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Sayısal sütunları formatla
            if col_num in [4, 6, 7]:  # Miktar ve tutar sütunları
                cell.number_format = '#,##0.00'
        
        # ADET sütunu hesaplama (9. sütun)
        unit = row_data[5]  # Birim
        quantity = row_data[4]  # Miktar
        
        adet_cell = ws_lines.cell(row=row_num, column=9)
        adet_cell.border = border
        
        if quantity is not None:
            if unit == 'TNE':
                # TNE ise: miktar * 1000 / 35
                adet_value = quantity * 1000 / 35
                adet_cell.value = adet_value
                adet_cell.number_format = '#,##0.00'
            elif unit == 'EA':
                # EA ise: direkt miktar değeri
                adet_cell.value = quantity
                adet_cell.number_format = '#,##0.00'
    
    # Sütun genişliklerini ayarla
    column_widths_lines = [10, 20, 10, 40, 12, 10, 15, 15, 15]
    for i, width in enumerate(column_widths_lines, 1):
        ws_lines.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # ========== İRSALİYELER SAYFAsl ==========
    ws_despatch = wb.create_sheet("İrsaliyeler")
    
    headers_despatch = [
        'Firma', 'Fatura No', 'İrsaliye No (Kısa)', 'İrsaliye No (Tam)', 
        'Tarih', 'Açıklama', 'Toplam Tutar (TL)'
    ]
    
    for col_num, header in enumerate(headers_despatch, 1):
        cell = ws_despatch.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    cursor.execute('''
        SELECT 
            i.firma_kodu, i.invoice_number, d.despatch_id_short, d.despatch_id_full,
            d.issue_date, d.description, i.total_amount
        FROM despatch_documents d
        JOIN invoices i ON d.invoice_id = i.id
        ORDER BY i.firma_kodu, i.invoice_number, d.despatch_id_short
    ''')
    
    for row_num, row_data in enumerate(cursor.fetchall(), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_despatch.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Toplam Tutar sütununu formatla (7. sütun)
            if col_num == 7:
                cell.number_format = '#,##0.00'
    
    # Sütun genişliklerini ayarla
    column_widths_desp = [10, 20, 20, 25, 12, 50, 18]
    for i, width in enumerate(column_widths_desp, 1):
        ws_despatch.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # ========== ÖZET SAYFAsl ==========
    ws_summary = wb.create_sheet("Özet", 0)  # İlk sayfa yap
    
    # Başlık
    ws_summary['A1'] = 'E-FATURA VERİTABANI ÖZETİ'
    ws_summary['A1'].font = Font(bold=True, size=16, color="366092")
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f'Rapor Tarihi: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws_summary['A2'].font = Font(italic=True, size=10)
    ws_summary.merge_cells('A2:D2')
    
    # İstatistikler
    row = 4
    
    # Toplam fatura sayısı
    cursor.execute('SELECT COUNT(*) FROM invoices')
    invoice_count = cursor.fetchone()[0]
    ws_summary[f'A{row}'] = 'Toplam Fatura Sayısı:'
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = invoice_count
    row += 1
    
    # Toplam tutar
    cursor.execute('SELECT SUM(total_amount) FROM invoices')
    total_amount = cursor.fetchone()[0]
    ws_summary[f'A{row}'] = 'Toplam Fatura Tutarı:'
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = f'{total_amount:,.2f} TRY'
    row += 1
    
    # Toplam vergi
    cursor.execute('SELECT SUM(tax_amount) FROM invoices')
    total_tax = cursor.fetchone()[0]
    ws_summary[f'A{row}'] = 'Toplam KDV Tutarı:'
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = f'{total_tax:,.2f} TRY'
    row += 1
    
    # Toplam satır sayısı
    cursor.execute('SELECT COUNT(*) FROM invoice_lines')
    line_count = cursor.fetchone()[0]
    ws_summary[f'A{row}'] = 'Toplam Fatura Satırı:'
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = line_count
    row += 1
    
    
    # Toplam irsaliye sayısı
    cursor.execute('SELECT COUNT(*) FROM despatch_documents')
    despatch_count = cursor.fetchone()[0]
    ws_summary[f'A{row}'] = 'Toplam İrsaliye:'
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = despatch_count
    row += 2
    
    # Fatura listesi (DISTINCT ile tekrarları engelle)
    ws_summary[f'A{row}'] = 'FATURA LİSTESİ'
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    cursor.execute('''
        SELECT DISTINCT invoice_number, issue_date, total_amount, supplier_name
        FROM invoices
        GROUP BY invoice_number
        ORDER BY issue_date DESC
    ''')
    
    for invoice in cursor.fetchall():
        ws_summary[f'A{row}'] = invoice[0]
        ws_summary[f'B{row}'] = invoice[1]
        ws_summary[f'C{row}'] = f'{invoice[2]:,.2f} TRY'
        ws_summary[f'D{row}'] = invoice[3]
        row += 1
    
    # Sütun genişliklerini ayarla
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 50
    
    # İstatistikleri al (connection kapatılmadan önce)
    cursor.execute("SELECT COUNT(*) FROM despatch_documents")
    despatch_count = cursor.fetchone()[0]
    
    conn.close()
    
    # Sabit dosya adı kullan (tek dosya)
    filename = 'data/excel/birlesik/efatura_birlesik.xlsx'
    
    # Eski dosya varsa sil
    if os.path.exists(filename):
        os.remove(filename)
        print(f"🗑️  Eski Excel dosyası silindi")
    
    # Yeni dosyayı kaydet
    wb.save(filename)
    
    print(f"\n{'=' * 70}")
    print(f"✓ Excel dosyası oluşturuldu: {filename}")
    print(f"{'=' * 70}")
    print(f"\nİçerik:")
    print(f"  📊 Özet sayfası")
    print(f"  📄 {invoice_count} fatura")
    print(f"  📋 {line_count} fatura satırı")
    print(f"  📦 {despatch_count} irsaliye")
    print(f"\nToplam Tutar: {total_amount:,.2f} TRY")
    print(f"Toplam KDV: {total_tax:,.2f} TRY")
    print()
    
    return filename

if __name__ == '__main__':
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl kütüphanesi bulunamadı. Yükleniyor...")
        import subprocess
        subprocess.check_call(['pip3', 'install', 'openpyxl'])
        import openpyxl
    
    create_excel_export()

